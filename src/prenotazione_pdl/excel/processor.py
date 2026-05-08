"""Modulo per l'elaborazione dei file Excel di configurazione e dati PDL."""

import getpass
import os
import time
from datetime import UTC, datetime
from typing import Any

import openpyxl
from loguru import logger
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

from ..config import Config
from ..models import CriticalConfigError, PDLData

# Gestione opzionale win32com per macro Excel (solo Windows)
WIN32COM_AVAILABLE = False
try:
    import win32com.client

    WIN32COM_AVAILABLE = True
except ImportError:
    pass


class ExcelProcessor:
    """Gestisce la lettura dei parametri da Excel e l'estrazione dei dati PDL."""

    def __init__(self, config_file_path: str, data_file_path: str | None = None) -> None:
        """
        Inizializza il processore Excel.

        Args:
            config_file_path: Percorso del file Excel di configurazione principale.
            data_file_path: Percorso del file dati PDL (se già noto).
        """
        self.config_file_path = config_file_path
        self.data_file_path = data_file_path
        logger.info(f"ExcelProcessor inizializzato. Config: '{config_file_path}'")
        if not os.path.exists(self.config_file_path):
            raise CriticalConfigError(f"File parametri Excel non trovato: {self.config_file_path}")

    @retry(
        stop=stop_after_attempt(Config.RETRY_GENERAL_ATTEMPTS),
        wait=wait_exponential(
            multiplier=Config.RETRY_WAIT_MULTIPLIER, min=Config.RETRY_WAIT_MIN_SECONDS, max=Config.RETRY_WAIT_MAX_SECONDS
        ),
        retry=retry_if_exception_type(Exception),
        reraise=True,
    )
    def _leggi_cella(self, file_path: str, sheet_name: str, cell_address: str, data_only: bool = True) -> Any:
        """Legge il valore di una singola cella con logica di retry."""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=data_only)
            sheet = wb[sheet_name]
            val = sheet[cell_address].value
            wb.close()
            return str(val).strip() if isinstance(val, str) else val
        except Exception as e:
            logger.error(f"Errore lettura cella {cell_address} in {file_path}: {e}")
            raise

    def get_website_url(self) -> str:
        """Recupera l'URL del sito SafeWork dal file di configurazione."""
        url = self._leggi_cella(self.config_file_path, Config.EXCEL_SHEET_PERCORSI, Config.CELLA_URL_SITO)
        if not isinstance(url, str) or not url.startswith("http"):
            logger.warning(f"URL non valido in Excel ('{url}'). Uso fallback: {Config.DEFAULT_URL_SITO}")
            return Config.DEFAULT_URL_SITO
        return url

    def get_pdl_data_file_path(self) -> str:
        """Recupera il percorso del file dati PDL dal file di configurazione."""
        path = self._leggi_cella(self.config_file_path, Config.EXCEL_SHEET_PERCORSI, Config.CELLA_PERCORSO_FILE_DATI_PDL)
        if not isinstance(path, str):
            raise CriticalConfigError("Percorso file dati PdL non trovato o non valido in Excel.")

        self.data_file_path = os.path.normpath(path)
        if not os.path.isabs(self.data_file_path):
            # Se è relativo, lo consideriamo relativo alla cartella dello script
            self.data_file_path = os.path.join(Config.SCRIPT_DIR, self.data_file_path)

        logger.info(f"Percorso file dati PdL risolto: {self.data_file_path}")
        return self.data_file_path

    def get_credentials(self, interactive_pwd: bool = False) -> tuple[str, str]:
        """Recupera username e password."""
        username = self._leggi_cella(self.config_file_path, Config.EXCEL_SHEET_CREDENTIALS, Config.USERNAME_CELL_EXCEL)
        if not username:
            raise CriticalConfigError("Username non trovato nel file di configurazione.")

        if interactive_pwd:
            password = getpass.getpass(f"Inserire password per '{username}': ")
        else:
            password = self._leggi_cella(
                self.config_file_path, Config.EXCEL_SHEET_CREDENTIALS, Config.PASSWORD_CELL_EXCEL
            )

        if not password:
            raise CriticalConfigError(f"Password non trovata per l'utente '{username}'.")

        return str(username), str(password)

    def run_pdl_macros(self) -> bool:
        """Esegue la sequenza di macro VBA necessaria per aggiornare i dati."""
        if not WIN32COM_AVAILABLE:
            logger.warning("Libreria pywin32 non disponibile. Esecuzione macro saltata.")
            return False

        if not self.data_file_path:
            self.get_pdl_data_file_path()

        if not self.data_file_path or not os.path.exists(self.data_file_path):
            logger.error(f"Impossibile eseguire macro: file '{self.data_file_path}' non trovato.")
            return False

        return self._esegui_sessione_macro(os.path.abspath(self.data_file_path))

    def _esegui_sessione_macro(self, abs_path: str) -> bool:
        """Gestisce una sessione Excel per l'esecuzione delle macro sequenziali."""
        excel_app, workbook = None, None
        success = False
        try:
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible, excel_app.DisplayAlerts = False, False
            workbook = excel_app.Workbooks.Open(abs_path)
            time.sleep(Config.PAUSA_APERTURA_EXCEL)

            for macro_name in [Config.MACRO_SEQ_1, Config.MACRO_SEQ_2, Config.MACRO_SEQ_3]:
                logger.info(f"Esecuzione macro '{macro_name}'...")
                excel_app.Application.Run(macro_name)
                time.sleep(1)

            workbook.Save()
            success = True
            logger.info("Tutte le macro sono state eseguite con successo.")
        except Exception as e:
            logger.error(f"Errore critico durante la sequenza macro: {e}", exc_info=True)
            raise
        finally:
            if workbook:
                workbook.Close(SaveChanges=success)
            if excel_app:
                excel_app.Quit()
        return success

    def get_pdl_list_from_excel(self) -> list[PDLData]:
        """Estrae la lista dei PDL da processare in un'unica lettura bulk del foglio."""
        if not self.data_file_path or not os.path.exists(self.data_file_path):
            return []

        colonna_giorno = self._get_colonna_giorno_corrente()
        if not colonna_giorno:
            return []

        logger.info(f"Estrazione PDL (Colonna Excel: {colonna_giorno})")
        return self._leggi_lista_bulk(colonna_giorno)

    def _get_colonna_giorno_corrente(self) -> str | None:
        """Determina la colonna Excel in base al giorno della settimana corrente."""
        oggi = datetime.now(UTC).date()
        giorno_settimana = oggi.weekday()
        if giorno_settimana not in Config.MAPPA_GIORNI_COLONNE_DATE:
            logger.warning(f"Oggi ({oggi.strftime('%A')}) non è un giorno previsto per l'automazione.")
            return None
        return Config.MAPPA_GIORNI_COLONNE_DATE[giorno_settimana]

    def _leggi_lista_bulk(self, colonna_giorno: str) -> list[PDLData]:
        """Esegue la lettura bulk delle righe nel foglio Excel."""
        lista_pdl: list[PDLData] = []
        try:
            wb = openpyxl.load_workbook(self.data_file_path, data_only=True, read_only=True)
            sheet = wb[Config.NOME_FOGLIO_DATI_PDL]
            idx_giorno = openpyxl.utils.column_index_from_string(colonna_giorno)
            col_map = self._genera_col_map()

            for i, row in enumerate(
                sheet.iter_rows(min_row=Config.RIGA_INIZIO_DATI_PDL, values_only=True), start=Config.RIGA_INIZIO_DATI_PDL
            ):
                pdl = self._processa_singola_riga(row, i, idx_giorno, col_map)
                if pdl:
                    lista_pdl.append(pdl)
            wb.close()
        except Exception as e:
            logger.error(f"Errore durante la lettura bulk della lista PDL: {e}", exc_info=True)
        return lista_pdl

    def _genera_col_map(self) -> dict[str, int]:
        """Genera la mappatura nome_colonna -> indice_colonna."""
        return {key: openpyxl.utils.column_index_from_string(col) for key, col in Config.COLONNE_EXCEL_REPORT.items()}

    def _processa_singola_riga(self, row: tuple[Any, ...], i: int, idx_giorno: int, col_map: dict[str, int]) -> PDLData | None:
        """Controlla il marker e processa la riga se necessario."""
        valore_giorno = row[idx_giorno - 1]
        if valore_giorno and str(valore_giorno).strip().upper() == 'X':
            pdl_info = self._processa_riga_vettoriale(row, i, col_map)
            return pdl_info if pdl_info.pdl else None
        return None

    def _processa_riga_vettoriale(self, row: tuple[Any, ...], riga_idx: int, col_map: dict[str, int]) -> PDLData:
        """Estrae i dati da una riga Excel e li incapsula in PDLData."""
        data: dict[str, Any] = {"riga_excel_debug": riga_idx}
        for key, col_idx in col_map.items():
            val = row[col_idx - 1]
            data[key] = self._formatta_valore_cella(val, key)
        return PDLData(**data)

    def _formatta_valore_cella(self, val: Any, key: str) -> str:
        """Uniforma il formato dei valori letti da Excel."""
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        if isinstance(val, (int, float)) and key == 'pdl':
            return str(int(val))
        return str(val).strip() if val is not None else ""
