# Importazioni necessarie
import argparse
import builtins
import contextlib
import getpass
import json
import logging
import os
import shutil
import time
import traceback
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    TimeoutException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_exponential

# --- DETERMINAZIONE ROBUSTA DELLA DIRECTORY DELLO SCRIPT ---
SCRIPT_FILE_PATH_ABS = os.path.abspath(__file__)
ACTUAL_SCRIPT_DIRECTORY = os.path.dirname(SCRIPT_FILE_PATH_ABS)
# --- FINE DETERMINAZIONE ROBUSTA ---

# --- Gestione Importazione win32com ---
WIN32COM_AVAILABLE = False
try:
    import win32com.client

    WIN32COM_AVAILABLE = True
except ImportError:
    pass

# --- Configurazione del Logging ---
root_logger = logging.getLogger()
for handler in root_logger.handlers[:]:
    root_logger.removeHandler(handler)
log_file_path = os.path.join(ACTUAL_SCRIPT_DIRECTORY, "automazione_pdl.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)-8s - %(name)-25s - %(funcName)-25s - %(message)s",
    handlers=[logging.FileHandler(log_file_path, mode="w", encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger("PDLAutomator")

if not WIN32COM_AVAILABLE:
    logger.critical(
        "!!! AVVISO CRITICO: La libreria 'pywin32' non è installata. "
        "Eseguire 'pip install pywin32'. Le funzionalità legate a Excel (macro) non funzioneranno."
    )


# --- Classe di Configurazione Interna ---
class Config:
    SCRIPT_DIR = ACTUAL_SCRIPT_DIRECTORY
    FILE_STATO_PROCESSO = "stato_processo_pdl.json"
    DEFAULT_URL_SITO = "https://safework.isab.com/"
    EXCEL_FILE_CONFIG_NAME = "parametri prenotazione pdl.xlsx"
    NOME_FOGLIO_DATI_PDL = "Riepilogo"
    EXCEL_SHEET_CREDENTIALS = "credenziali"
    EXCEL_SHEET_PERCORSI = "percorsi"
    CELLA_URL_SITO = "B3"
    CELLA_PERCORSO_FILE_DATI_PDL = "B2"
    USERNAME_CELL_EXCEL = "A3"
    PASSWORD_CELL_EXCEL = "B3"
    COLONNE_EXCEL_REPORT = {
        "pdl": "E",
        "area": "D",
        "descrizione": "G",
        "stato_pdl_excel": "M",
        "stato_attivita_excel": "Q",
        "data_controllo_excel": "R",
        "personale_excel": "S",
        "impianto": "F",  # Aggiunta colonna IMPIANTO
    }
    RIGA_INIZIO_DATI_PDL = 4
    MAPPA_GIORNI_COLONNE_DATE = {i: chr(ord("H") + i) for i in range(5)}
    MACRO_SEQ_1 = "ordina.OrdinaEFormattaTabellaCorrente"
    MACRO_SEQ_2 = "Modulo8.aggiornaQuery"
    MACRO_SEQ_3 = "Rimuovi_Tutti_I_Filtri.RimuoviTuttiIFiltri"
    DROPDOWN_SITO_SELECTORS = [(By.XPATH, "//button[@class='ms-choice']")]
    ISAB_SUD_OPTION_SELECTORS = [
        (By.XPATH, "//div[contains(@class, 'ms-drop')]//label[.//span[normalize-space()='ISAB Sud']]")
    ]
    USERNAME_FIELD_SELECTORS = [(By.ID, "inpUtente")]
    PASSWORD_FIELD_SELECTORS = [(By.ID, "inpPassword")]
    LOGIN_BUTTON_SELECTORS = [(By.ID, "btnLogin")]
    INDICATORE_CARICAMENTO_LOGIN_SELECTORS = [(By.XPATH, "//*[contains(text(), 'Caricamento...')]")]
    HOME_BUTTON_SELECTORS = [(By.ID, "topIcon-actHomePage")]
    LINK_PRENOTAZIONE_PDL_SELECTORS = [(By.ID, "sideBar-actPrenotazionePdL")]
    INPUT_PDL_WEB_SELECTORS = [(By.ID, "inpNumPermessoApparecchitura")]
    TASTO_CERCA_PDL_SELECTORS = [(By.ID, "btnAvviaRicerca")]
    ICON_DA_PRENOTARE_SELECTORS = [
        (By.XPATH, "//i[@title='Da prenotare' and contains(@class, 'daprenotare')]")
    ]
    ICON_GIA_PRENOTATO_SELECTORS = [(By.XPATH, "//i[@title='Prenotato' and contains(@class, 'prenotato')]")]
    TASTO_SALVA_PRENOTAZIONE_SELECTORS = [(By.ID, "btnSalva")]
    MSG_PDL_NON_TROVATO_SELECTORS = [
        (
            By.XPATH,
            "//span[normalize-space(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZÀÈÉÌÒÙ', 'abcdefghijklmnopqrstuvwxyzàèéìòù'))='nessun permesso trovato']",
        )
    ]
    TIMEOUT_ALERT_P_TAG_SELECTORS = [
        (By.XPATH, "//p[@idtxt='CD69E8EA'][contains(text(), 'Request timeout (30000ms)')]")
    ]
    TIMEOUT_ALERT_FALLBACK_SELECTORS = [
        (
            By.XPATH,
            "//div[contains(@class, 'bootstrap-dialog-message')]//p[contains(text(), 'Request timeout (30000ms)')]",
        )
    ]
    TIMEOUT_ALERT_OK_BUTTON_SELECTORS = [
        (
            By.XPATH,
            "//div[contains(@class, 'bootstrap-dialog') or contains(@class, 'modal-dialog')]//button[normalize-space(translate(., 'OKok', 'okok'))='ok']",
        )
    ]
    RETRY_GENERAL_ATTEMPTS = 3
    RETRY_WAIT_MULTIPLIER = 1
    RETRY_WAIT_MIN_SECONDS = 2
    RETRY_WAIT_MAX_SECONDS = 10
    MAX_SETUP_ATTEMPTS = 5
    SELENIUM_TIMEOUT_PAGE_LOAD = 90
    SELENIUM_TIMEOUT_LONG = 120
    SELENIUM_TIMEOUT_MEDIUM = 40
    SELENIUM_TIMEOUT_SHORT = 20
    SELENIUM_TIMEOUT_QUICK_CHECK = 5
    PAUSE_GENERAL_SHORT = 0.5
    PAUSE_GENERAL_MEDIUM = 1.5
    PAUSA_DOPO_GET_INIZIALE = 2
    PAUSA_DOPO_CLICK_DROPDOWN_SITO = 1.5
    PAUSA_DOPO_SELEZIONE_SITO = 1
    PAUSA_DOPO_CLICK_DA_PRENOTARE = 1.0
    PAUSA_TRA_PDL = 1.0
    PAUSA_TRA_TENTATIVI_SETUP_FALLITI = 3
    PAUSA_APERTURA_EXCEL = 3
    PAUSA_ATTIVAZIONE_FOGLIO_EXCEL = 1
    PAUSA_COMPLETAMENTO_MACRO = 10


# --- Custom Exceptions, Dataclass, WebDriverManager, ExcelProcessor ---
class CriticalConfigError(Exception):
    pass


class ProcessInterruptedException(Exception):
    pass


class AutomationException(Exception):
    pass


class TimeoutAlertDetected(AutomationException):
    pass


@dataclass
class PDLData:
    riga_excel_debug: int
    pdl: str = ""
    area: str = "Area Non Specificata"
    descrizione: str = ""
    stato_pdl_excel: str = ""
    stato_attivita_excel: str = ""
    data_controllo_excel: str = ""
    personale_excel: str = ""
    impianto: str = ""
    stato_script: str = "Non Processato"  # Aggiunto impianto


class WebDriverManager:
    def __init__(self, headless: bool = True, start_maximized: bool = False) -> None:
        self.headless = headless
        self.start_maximized = start_maximized
        self.driver: webdriver.Chrome | None = None
        logger.info(f"WebDriverManager (headless={headless}, start_maximized={start_maximized})")

    def get_driver(self) -> webdriver.Chrome:
        if self.driver and self._is_driver_alive():
            logger.debug("Restituisco driver esistente.")
            return self.driver
        logger.info("Creazione WebDriver Chrome.")
        options = Options()
        if self.headless:
            options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--log-level=3")
        options.add_argument("--disable-features=PasswordLeakDetection")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("enable-automation")
        options.add_argument("--disable-infobars")
        options.add_argument("--disable-browser-side-navigation")
        if self.start_maximized and not self.headless:
            options.add_argument("--start-maximized")
        else:
            options.add_argument(
                f"--window-size={Config.RETRY_WAIT_MAX_SECONDS * 148},{Config.RETRY_WAIT_MAX_SECONDS * 102 + 4}"
            )

        prefs = {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.password_manager_leak_detection": False,
        }
        options.add_experimental_option("prefs", prefs)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])

        try:
            self.driver = webdriver.Chrome(options=options)
            self.driver.set_page_load_timeout(Config.SELENIUM_TIMEOUT_PAGE_LOAD)
            logger.info("WebDriver Chrome creato.")
        except Exception as e:
            logger.error(f"Errore creazione WebDriver: {e}", exc_info=True)
            raise AutomationException(f"Impossibile creare WebDriver: {e}")
        return self.driver

    def _is_driver_alive(self) -> bool:
        if not self.driver:
            return False
        try:
            _ = self.driver.title
            return True
        except:
            return False

    def quit_driver(self) -> None:
        if self.driver:
            try:
                logger.info("Chiusura WebDriver.")
                self.driver.quit()
            except Exception as e:
                logger.warning(f"Errore chiusura WebDriver: {e}", exc_info=False)
            finally:
                self.driver = None

    def restart_driver(self) -> webdriver.Chrome:
        logger.info("Riavvio WebDriver.")
        self.quit_driver()
        time.sleep(Config.PAUSE_GENERAL_MEDIUM)
        return self.get_driver()


class ExcelProcessor:
    def __init__(self, config_file_path: str, data_file_path: str) -> None:
        self.config_file_path = config_file_path
        self.data_file_path = data_file_path
        logger.info(
            f"ExcelProcessor. Config: '{config_file_path}', Data: '{data_file_path or '<NON ANCORA DEFINITO>'}'"
        )
        if not os.path.exists(self.config_file_path):
            raise CriticalConfigError(f"File parametri Excel non trovato: {self.config_file_path}")

    @retry(
        stop=stop_after_attempt(Config.RETRY_GENERAL_ATTEMPTS),
        wait=wait_exponential(
            multiplier=Config.RETRY_WAIT_MULTIPLIER,
            min=Config.RETRY_WAIT_MIN_SECONDS,
            max=Config.RETRY_WAIT_MAX_SECONDS,
        ),
        retry=retry_if_exception_type(Exception),
        reraise=True,
    )
    def _leggi_cella(self, fp: str, sn: str, ca: str, d_o: bool = True) -> Any:
        try:
            wb = openpyxl.load_workbook(fp, data_only=d_o)
            sh = wb[sn]
            val = sh[ca].value
            wb.close()
            return str(val).strip() if isinstance(val, str) else val
        except Exception as e:
            logger.error(f"Errore lettura {ca} da {fp},'{sn}': {e}", exc_info=True)
            raise

    def get_website_url(self) -> str:
        url = self._leggi_cella(self.config_file_path, Config.EXCEL_SHEET_PERCORSI, Config.CELLA_URL_SITO)
        if not url or not isinstance(url, str) or not url.startswith("http"):
            logger.warning(f"URL non valido ('{url}'). Fallback: {Config.DEFAULT_URL_SITO}")
            return Config.DEFAULT_URL_SITO
        logger.info(f"URL sito: {url}")
        return url

    def get_pdl_data_file_path(self) -> str:
        path = self._leggi_cella(
            self.config_file_path, Config.EXCEL_SHEET_PERCORSI, Config.CELLA_PERCORSO_FILE_DATI_PDL
        )
        if not path or not isinstance(path, str):
            raise CriticalConfigError("Percorso file dati PdL non valido.")
        self.data_file_path = os.path.normpath(path)
        logger.info(f"Percorso file dati PdL: {self.data_file_path}")
        if not os.path.exists(self.data_file_path):
            logger.warning(f"File dati PdL '{self.data_file_path}' non esiste.")
        return self.data_file_path

    def get_credentials(self, secure_pwd_in: bool) -> tuple[str, str]:
        usr = self._leggi_cella(
            self.config_file_path, Config.EXCEL_SHEET_CREDENTIALS, Config.USERNAME_CELL_EXCEL
        )
        if not usr:
            raise CriticalConfigError("Username non trovato.")
        pwd = (
            getpass.getpass(f"PW per '{usr}': ")
            if secure_pwd_in
            else self._leggi_cella(
                self.config_file_path, Config.EXCEL_SHEET_CREDENTIALS, Config.PASSWORD_CELL_EXCEL
            )
        )
        if not secure_pwd_in:
            logger.warning("LETTURA PW DA EXCEL: NON SICURO!")
        if not pwd:
            raise CriticalConfigError("Password non fornita.")
        logger.info(f"Credenziali per '{usr}' (PW non loggata).")
        return usr, pwd

    @retry(
        stop=stop_after_attempt(Config.RETRY_GENERAL_ATTEMPTS - 1),
        wait=wait_exponential(multiplier=1, min=5, max=20),
        retry=retry_if_exception_type(Exception),
        reraise=True,
    )
    def _esegui_macro(self, nm: str, nfd: str | None = None, fp_o: str | None = None) -> bool:
        if not WIN32COM_AVAILABLE:
            raise ImportError("win32com.client non disponibile per macro.")
        f_proc = fp_o or self.data_file_path
        if not f_proc or not os.path.exists(f_proc):
            logger.error(f"File Excel '{f_proc}' non trovato per macro '{nm}'.")
            return False
        abs_p = os.path.abspath(f_proc)
        ex = None
        wb = None
        ok = False
        logger.info(f"Esecuzione macro '{nm}' su '{abs_p}'.")
        try:
            ex = win32com.client.Dispatch("Excel.Application")
            ex.Visible = False
            ex.DisplayAlerts = False
            wb = ex.Workbooks.Open(abs_p)
            time.sleep(Config.PAUSA_APERTURA_EXCEL)
            if nfd:
                try:
                    sh = wb.Sheets(nfd)
                    sh.Activate()
                    logger.info(f"Foglio '{nfd}' attivato.")
                    time.sleep(Config.PAUSA_ATTIVAZIONE_FOGLIO_EXCEL)
                except Exception as esh:
                    logger.warning(f"Impossibile attivare foglio '{nfd}': {esh}")
            logger.info(f"Esecuzione Application.Run('{nm}').")
            ex.Application.Run(nm)
            logger.info(f"Macro '{nm}' avviata. Pausa {Config.PAUSA_COMPLETAMENTO_MACRO}s.")
            time.sleep(Config.PAUSA_COMPLETAMENTO_MACRO)
            logger.info("Salvataggio workbook...")
            wb.Save()
            logger.info(f"Macro '{nm}' eseguita e file salvato.")
            ok = True
            return True
        except Exception as e:
            logger.error(f"Errore durante esecuzione macro '{nm}': {e}", exc_info=True)
            raise
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=ok)
                if ex is not None:
                    ex.Quit()
            except Exception as ec:
                logger.warning(f"Errore chiusura Excel per macro '{nm}': {ec}", exc_info=False)
        return False

    def run_pdl_macros(self) -> bool:
        logger.info("Esecuzione sequenza macro PdL.")
        if not self.data_file_path:
            self.get_pdl_data_file_path()
        if not self.data_file_path or not os.path.exists(self.data_file_path):
            logger.error(f"File dati '{self.data_file_path}' non trovato.")
            return False
        macros = [
            (Config.MACRO_SEQ_1, Config.NOME_FOGLIO_DATI_PDL),
            (Config.MACRO_SEQ_2, Config.NOME_FOGLIO_DATI_PDL),
            (Config.MACRO_SEQ_3, Config.NOME_FOGLIO_DATI_PDL),
        ]
        for nm, nf in macros:
            if not self._esegui_macro(nm, nf):
                logger.error(f"Macro '{nm}' fallita.")
                return False
            logger.info(f"Macro '{nm}' OK.")
        logger.info("Tutte macro PdL OK.")
        return True

    def get_pdl_list_from_excel(self) -> list[PDLData]:
        logger.info(f"Lettura lista PdL da Excel: '{Config.NOME_FOGLIO_DATI_PDL}'.")
        if not self.data_file_path or not os.path.exists(self.data_file_path):
            logger.error(f"File dati '{self.data_file_path}' non trovato.")
            return []
        lista_pdl = []
        oggi = date.today()
        gio_set = oggi.weekday()
        if gio_set not in Config.MAPPA_GIORNI_COLONNE_DATE:
            logger.warning(f"Oggi ({oggi.strftime('%A')}) non mappato.")
            return []
        col_gio = Config.MAPPA_GIORNI_COLONNE_DATE[gio_set]
        logger.info(f"Processo PdL per {oggi.strftime('%d/%m/%Y')}, col: {col_gio}")
        try:
            wb = openpyxl.load_workbook(self.data_file_path, data_only=True)
            sh = wb[Config.NOME_FOGLIO_DATI_PDL]
            for r_n in range(Config.RIGA_INIZIO_DATI_PDL, sh.max_row + 1):
                try:
                    if str(sh[f"{col_gio}{r_n}"].value).strip().upper() == "X":
                        p_d: dict[str, Any] = {"riga_excel_debug": r_n}
                        for k, cl in Config.COLONNE_EXCEL_REPORT.items():
                            v = sh[f"{cl}{r_n}"].value
                            if isinstance(v, datetime):
                                p_d[k] = v.strftime("%d/%m/%Y")
                            elif isinstance(v, (int, float)) and k == "pdl":
                                p_d[k] = str(int(v))
                            elif v is not None:
                                p_d[k] = str(v).strip()
                            else:
                                p_d[k] = ""
                        if p_d.get("pdl"):
                            lista_pdl.append(PDLData(**p_d))
                        else:
                            logger.warning(f"Riga {r_n}: PdL mancante con 'X'.")
                except Exception as er:
                    logger.exception(f"Errore riga {r_n}: {er}", exc_info=False)
            wb.close()
        except Exception as e:
            logger.error(f"Errore lettura lista PdL: {e}", exc_info=True)
            return []
        logger.info(f"Ottenuti {len(lista_pdl)} PdL." if lista_pdl else "Nessun PdL.")
        return lista_pdl


# --- SafeWork Automator ---
class SafeWorkAutomator:
    def __init__(self, driver: webdriver.Chrome, dry_run: bool = False) -> None:
        self.driver = driver
        self.dry_run = dry_run
        self.current_pdl_for_alert_context: str | None = None
        logger.info(f"SafeWorkAutomator (dry_run={dry_run})")

    def _attendi_caricamento_pagina(
        self, timeout_overlay: int = Config.SELENIUM_TIMEOUT_LONG, timeout_popup: int = 3
    ) -> None:
        logger.info("Attesa overlay 'GISWaitOverlay' e popup generici...")
        try:
            wait_overlay = WebDriverWait(self.driver, timeout_overlay)
            wait_overlay.until(EC.invisibility_of_element_located((By.ID, "GISWaitOverlay")))
            logger.info(" -> Overlay 'GISWaitOverlay' scomparso.")
        except TimeoutException:
            logger.warning(f"Timeout ({timeout_overlay}s) attesa 'GISWaitOverlay'. Proseguo con cautela.")
        try:
            wait_popup = WebDriverWait(self.driver, timeout_popup)
            modal_attivo = wait_popup.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//div[contains(@class, 'modal') and contains(@style, 'display: block')]")
                )
            )
            logger.warning("Trovato un popup modale imprevisto. Tento di chiuderlo.")
            if not self.dry_run:
                bottone_chiusura = modal_attivo.find_element(
                    By.XPATH, ".//button[contains(text(), 'OK') or @data-dismiss='modal']"
                )
                bottone_chiusura.click()
                logger.info(" -> Popup modale chiuso.")
                WebDriverWait(self.driver, 10).until(EC.invisibility_of_element(modal_attivo))
            else:
                logger.info("[DRY RUN] Simulata chiusura popup modale.")
        except TimeoutException:
            pass
        except Exception as e:
            logger.exception(f"Impossibile chiudere popup modale: {e}")
        time.sleep(Config.PAUSE_GENERAL_SHORT)
        logger.info("Pagina stabilizzata (overlay/popup).")

    def _attendi_indicatore_login(self, timeout: int = Config.SELENIUM_TIMEOUT_LONG) -> None:
        logger.info("Attesa indicatore testuale 'Caricamento...' dopo il login.")
        try:
            wait = WebDriverWait(self.driver, timeout)
            selettore_indicatore = Config.INDICATORE_CARICAMENTO_LOGIN_SELECTORS[0]
            wait.until(EC.visibility_of_element_located(selettore_indicatore))
            logger.info(" -> Indicatore 'Caricamento...' apparso.")
            wait.until(EC.invisibility_of_element_located(selettore_indicatore))
            logger.info(" -> Indicatore 'Caricamento...' scomparso.")
        except TimeoutException:
            logger.warning(
                f"Timeout ({timeout}s) durante attesa indicatore 'Caricamento...'. Proseguo ugualmente."
            )

    def _find_element(
        self,
        selectors: Sequence[tuple[By | str, str]],
        wait_time: float | None = None,
        condition: Any = EC.visibility_of_element_located,
    ) -> Any:
        wait_s = wait_time or Config.SELENIUM_TIMEOUT_MEDIUM
        wait = WebDriverWait(self.driver, wait_s)
        last_ex = None
        for strat, val in selectors:
            try:
                logger.debug(f"Trova elemento: {strat}='{val}', wait: {wait_s}s")
                el = wait.until(condition((strat, val)))
                logger.info(f"Elemento trovato: {strat}='{val}'")
                return el
            except TimeoutException as e:
                last_ex = e
                logger.warning(f"Timeout: {strat}='{val}'. Fallback...")
            except Exception as eg:
                last_ex: Any = eg
                logger.warning(f"Eccezione ({type(eg).__name__}): {strat}='{val}'. Fallback...")
        pdl_c = f" (PdL: {self.current_pdl_for_alert_context})" if self.current_pdl_for_alert_context else ""
        logger.error(f"Elemento non trovato con {selectors}{pdl_c}.")
        if last_ex:
            raise AutomationException(f"Elemento non trovato con {selectors}: {last_ex}") from last_ex
        raise NoSuchElementException(f"Elemento non trovato: {selectors}")

    def _click_element(
        self, selectors: Sequence[tuple[By | str, str]], el_name: str, wait_time: float | None = None
    ) -> None:
        logger.info(f"Click su: '{el_name}'")
        if self.dry_run:
            logger.info(f"[DRY RUN] Click su '{el_name}' con {selectors}.")
            return
        el = self._find_element(selectors, wait_time, EC.element_to_be_clickable)
        try:
            self.driver.execute_script("arguments[0].click();", el)
            logger.info(f"Click JS su '{el_name}'.")
        except Exception as ejs:
            logger.warning(f"Click JS su '{el_name}' fallito ({ejs}), tento standard.")
            try:
                el.click()
                logger.info(f"Click standard su '{el_name}'.")
            except ElementClickInterceptedException:
                logger.warning(f"Click intercettato su '{el_name}', provo ActionChains.", exc_info=False)
                try:
                    ActionChains(self.driver).move_to_element(el).click().perform()
                    logger.info(f"Click ActionChains su '{el_name}'.")
                except Exception as ea:
                    logger.error(f"Falliti tutti click per '{el_name}': {ea}", exc_info=True)
                    raise AutomationException(f"Click '{el_name}' fallito: {ea}") from ea
            except Exception as es:
                logger.error(f"Click standard su '{el_name}' fallito: {es}", exc_info=True)
                raise AutomationException(f"Click '{el_name}' fallito: {es}") from es

    def _pulisci_e_inserisci_testo(
        self, selectors: Sequence[tuple[By | str, str]], text: str, f_name: str
    ) -> None:
        logger.info(f"Inserimento '{text or '<VUOTO>'}' in '{f_name}'.")
        if self.dry_run:
            logger.info(f"[DRY RUN] Inserimento '{text}' in '{f_name}'.")
            return
        el = self._find_element(selectors, condition=EC.visibility_of_element_located)
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
            time.sleep(0.2)
            el.click()
            time.sleep(0.2)
            el.send_keys(Keys.CONTROL + "a")
            time.sleep(0.2)
            el.send_keys(Keys.BACK_SPACE)
            time.sleep(0.2)
            if text:
                el.send_keys(text)
            logger.info(f"Testo '{text or '<VUOTO>'}' inserito in '{f_name}'.")
        except Exception as e:
            logger.error(f"Fallito inserimento in '{f_name}': {e}", exc_info=True)
            raise AutomationException(f"Interazione con '{f_name}' fallita.") from e

    def check_and_handle_specific_timeout_alert(self, op_ctx: str) -> None:
        alert_els = []
        try:
            with contextlib.suppress(builtins.BaseException):
                alert_els = self.driver.find_elements(*Config.TIMEOUT_ALERT_P_TAG_SELECTORS[0])
            if not alert_els:
                with contextlib.suppress(builtins.BaseException):
                    alert_els = self.driver.find_elements(*Config.TIMEOUT_ALERT_FALLBACK_SELECTORS[0])
            for ae in alert_els:
                if ae.is_displayed():
                    alert_txt = ae.text
                    pdl_c = (
                        f" (PdL: {self.current_pdl_for_alert_context})"
                        if self.current_pdl_for_alert_context
                        else ""
                    )
                    logger.error(
                        f"!!! ALERT TIMEOUT SITO !!! Contesto: '{op_ctx}'{pdl_c}. Testo: '{alert_txt}'."
                    )
                    if not self.dry_run:
                        try:
                            ok_b = self._find_element(
                                Config.TIMEOUT_ALERT_OK_BUTTON_SELECTORS,
                                Config.SELENIUM_TIMEOUT_QUICK_CHECK,
                                EC.element_to_be_clickable,
                            )
                            logger.info("Click OK su alert timeout sito.")
                            ok_b.click()
                            time.sleep(Config.PAUSE_GENERAL_MEDIUM)
                        except Exception as e_ok:
                            logger.warning(f"Impossibile click OK su alert: {e_ok}", exc_info=False)
                    else:
                        logger.info("[DRY RUN] Simulata chiusura alert timeout sito.")
                    raise TimeoutAlertDetected(
                        f"Alert timeout sito '{alert_txt}' rilevato durante '{op_ctx}'{pdl_c}."
                    )
        except TimeoutAlertDetected:
            raise
        except Exception as ec:
            logger.debug(f"Errore minore check alert ({op_ctx}): {type(ec).__name__}", exc_info=False)

    @retry(
        stop=stop_after_attempt(Config.RETRY_GENERAL_ATTEMPTS),
        wait=wait_exponential(
            multiplier=Config.RETRY_WAIT_MULTIPLIER,
            min=Config.RETRY_WAIT_MIN_SECONDS,
            max=Config.RETRY_WAIT_MAX_SECONDS,
        ),
        retry=retry_if_exception_type(AutomationException),
        reraise=True,
    )
    def login(self, url: str, usr: str, pwd: str) -> None:
        logger.info(f"Login a: {url} con utente: {usr}")
        if self.dry_run:
            logger.info(f"[DRY RUN] Login a {url}")
            return
        self.driver.get(url)
        time.sleep(Config.PAUSA_DOPO_GET_INIZIALE)
        if "data:," in self.driver.current_url:
            logger.warning("Pagina 'data:,'. Riprovo GET.")
            self.driver.get(url)
            time.sleep(Config.PAUSA_DOPO_GET_INIZIALE * 2)
            if "data:," in self.driver.current_url:
                raise AutomationException(f"Fallimento URL: {url}. 'data:,' persistente.")
        try:
            WebDriverWait(self.driver, Config.SELENIUM_TIMEOUT_SHORT).until(lambda d: "SafeWork" in d.title)
            logger.info(f"Pagina login caricata. Titolo: '{self.driver.title}'")
        except TimeoutException:
            logger.exception(
                f"Timeout caricamento pagina login. URL:{self.driver.current_url}, Titolo:{self.driver.title}"
            )
            raise AutomationException("Timeout caricamento pagina login.")
        self._click_element(Config.DROPDOWN_SITO_SELECTORS, "Dropdown Sito")
        time.sleep(Config.PAUSA_DOPO_CLICK_DROPDOWN_SITO)
        self._click_element(Config.ISAB_SUD_OPTION_SELECTORS, "Opzione ISAB Sud")
        time.sleep(Config.PAUSA_DOPO_SELEZIONE_SITO)
        self._pulisci_e_inserisci_testo(Config.USERNAME_FIELD_SELECTORS, usr, "Username")
        self._pulisci_e_inserisci_testo(Config.PASSWORD_FIELD_SELECTORS, pwd, "Password")
        self._click_element(Config.LOGIN_BUTTON_SELECTORS, "Login Button")
        logger.info("Attesa della scomparsa del form di login (segnale di successo)...")
        try:
            wait = WebDriverWait(self.driver, Config.SELENIUM_TIMEOUT_MEDIUM)
            wait.until(EC.invisibility_of_element_located(Config.USERNAME_FIELD_SELECTORS[0]))
            logger.info(" -> Form di login scomparso. La pagina sta caricando.")
        except TimeoutException:
            logger.exception("Il form di login non è scomparso. Login probabilmente fallito.")
            raise AutomationException("Login fallito, form ancora presente.")
        self._attendi_indicatore_login()
        self._attendi_caricamento_pagina()
        self.check_and_handle_specific_timeout_alert("dopo invio login")
        logger.info("Login completato.")

    def navigate_to_pdl_booking_page(self) -> None:
        logger.info("Navigazione a pagina prenotazione PdL.")
        self._attendi_caricamento_pagina()
        self._click_element(Config.HOME_BUTTON_SELECTORS, "Pulsante Home")
        self._attendi_caricamento_pagina()
        self.check_and_handle_specific_timeout_alert("dopo click Home")
        el_prenotazione = self._find_element(
            Config.LINK_PRENOTAZIONE_PDL_SELECTORS, condition=EC.element_to_be_clickable
        )
        if not self.dry_run:
            self.driver.execute_script(
                "arguments[0].scrollIntoView({behavior:'auto',block:'center'});", el_prenotazione
            )
            time.sleep(Config.PAUSE_GENERAL_SHORT)
            el_prenotazione.click()
        else:
            logger.info("[DRY RUN] Click 'Prenotazione PdL'.")
        self._attendi_caricamento_pagina()
        self.check_and_handle_specific_timeout_alert("dopo click Prenotazione PdL")
        self._find_element(
            Config.INPUT_PDL_WEB_SELECTORS, Config.SELENIUM_TIMEOUT_MEDIUM, EC.element_to_be_clickable
        )
        logger.info("Pagina prenotazione PdL caricata.")

    def process_single_pdl(self, pdl_data: PDLData) -> str:
        self.current_pdl_for_alert_context = pdl_data.pdl
        logger.info(f"Processo PdL: '{pdl_data.pdl}' (Area: '{pdl_data.area}')")
        self._pulisci_e_inserisci_testo(
            Config.INPUT_PDL_WEB_SELECTORS, pdl_data.pdl, f"Input PdL ({pdl_data.pdl})"
        )
        self._click_element(Config.TASTO_CERCA_PDL_SELECTORS, "Tasto Cerca PdL")
        self._attendi_caricamento_pagina()
        self.check_and_handle_specific_timeout_alert(f"dopo click Cerca per PdL {pdl_data.pdl}")
        try:
            self._find_element(Config.MSG_PDL_NON_TROVATO_SELECTORS, Config.SELENIUM_TIMEOUT_QUICK_CHECK)
            logger.info(f"PdL '{pdl_data.pdl}': 'Nessun permesso trovato'.")
            self.current_pdl_for_alert_context = None
            return "Non Trovato (ricerca web)"
        except (AutomationException, NoSuchElementException, TimeoutException):
            logger.debug(f"PdL '{pdl_data.pdl}': 'Nessun permesso trovato' NON rilevato.")
        try:
            self._find_element(Config.ICON_GIA_PRENOTATO_SELECTORS, Config.SELENIUM_TIMEOUT_QUICK_CHECK)
            logger.info(f"PdL '{pdl_data.pdl}': 'Già Prenotato'.")
            self.current_pdl_for_alert_context = None
            return "Già Prenotato"
        except (AutomationException, NoSuchElementException, TimeoutException):
            logger.info(f"PdL '{pdl_data.pdl}': 'Già Prenotato' non rilevato.")
        logger.info(f"PdL '{pdl_data.pdl}' da prenotare...")
        try:
            self._click_element(Config.ICON_DA_PRENOTARE_SELECTORS, "Icona Da Prenotare")
            time.sleep(Config.PAUSA_DOPO_CLICK_DA_PRENOTARE)
            self._click_element(Config.TASTO_SALVA_PRENOTAZIONE_SELECTORS, "Tasto Salva Prenotazione")
            self._attendi_caricamento_pagina()
            logger.info(f"PdL '{pdl_data.pdl}': Prenotazione Eseguita.")
            self.current_pdl_for_alert_context = None
            return "Prenotazione Eseguita"
        except TimeoutAlertDetected:
            raise
        except Exception as e:
            logger.error(
                f"PdL '{pdl_data.pdl}': Errore azioni prenotazione: {type(e).__name__}", exc_info=True
            )
            pdl_data.stato_script = f"Fallimento ({type(e).__name__})"
        self.current_pdl_for_alert_context = None
        return pdl_data.stato_script


# --- StateManager, PDLOrchestrator e __main__ ---
class StateManager:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path
        logger.info(f"StateManager per: {file_path}")

    def carica_stato(self) -> tuple[int, list[PDLData]]:
        if os.path.exists(self.file_path):
            try:
                with open(self.file_path, encoding="utf-8") as f:
                    stato = json.load(f)
                idx = stato.get("ultimo_indice_pdl_processato", -1)
                res_d = stato.get("risultati_elaborazione_completi", [])
                res_pdl = [PDLData(**pd) for pd in res_d]
                logger.info(f"Stato caricato. Indice: {idx}. {len(res_pdl)} risultati.")
                return idx, res_pdl
            except Exception as e:
                logger.warning(
                    f"Errore caricamento stato '{self.file_path}': {e}. Si ricomincia.", exc_info=True
                )
                try:
                    shutil.move(
                        self.file_path,
                        self.file_path + f".corrupted_{datetime.now().strftime('%Y%m%d%H%M%S')}",
                    )
                except Exception as em:
                    logger.exception(f"Impossibile rinominare file stato corrotto: {em}")
        logger.info(f"File stato '{self.file_path}' non trovato/illeggibile. Partenza da zero.")
        return -1, []

    def salva_stato(self, idx: int, res: list[PDLData]) -> None:
        res_s = [p.__dict__ for p in res]
        stato_s = {"ultimo_indice_pdl_processato": idx, "risultati_elaborazione_completi": res_s}
        try:
            with open(self.file_path, "w", encoding="utf-8") as f:
                json.dump(stato_s, f, indent=4, ensure_ascii=False)
            logger.info(f"Stato salvato (indice: {idx}) in '{self.file_path}'.")
        except Exception as e:
            logger.error(f"Errore salvataggio stato '{self.file_path}': {e}", exc_info=True)

    def rimuovi_file_stato(self) -> None:
        if os.path.exists(self.file_path):
            try:
                os.remove(self.file_path)
                logger.info(f"File stato '{self.file_path}' rimosso.")
            except Exception as e:
                logger.error(f"Impossibile rimuovere file stato '{self.file_path}': {e}", exc_info=True)


class PDLOrchestrator:
    def __init__(self, dry_run: bool = False, secure_password_input: bool = False) -> None:
        self.dry_run = dry_run
        self.secure_password_input = secure_password_input
        logger.info(f"PDLOrchestrator (dry_run={dry_run}, secure_password={secure_password_input})")
        self.config_excel_path = os.path.join(Config.SCRIPT_DIR, Config.EXCEL_FILE_CONFIG_NAME)
        self.excel_processor = ExcelProcessor(self.config_excel_path, data_file_path="")
        self.pdl_data_file_path = self.excel_processor.get_pdl_data_file_path()
        self.web_driver_manager = WebDriverManager(headless=False, start_maximized=True)
        self.state_manager = StateManager(os.path.join(Config.SCRIPT_DIR, Config.FILE_STATO_PROCESSO))
        self.website_url: str = ""
        self.username: str = ""
        self.password: str = ""

    def _setup_credentials_and_url(self) -> None:
        logger.info("Setup credenziali e URL.")
        self.website_url = self.excel_processor.get_website_url()
        self.username, self.password = self.excel_processor.get_credentials(self.secure_password_input)
        if not all([self.website_url, self.username, self.password]):
            raise CriticalConfigError("URL/usr/pw mancanti.")

    def _initialize_pdl_list(self) -> list[PDLData]:
        logger.info("Inizializzazione lista PdL.")
        if not self.dry_run:
            if not self.excel_processor.run_pdl_macros():
                raise AutomationException("Fallimento macro Excel.")
        else:
            logger.info("[DRY RUN] Esecuzione macro saltata.")
        pdl_excel = self.excel_processor.get_pdl_list_from_excel()
        if not pdl_excel:
            logger.info("Nessun PdL da Excel dopo macro (o salto).")
        return pdl_excel

    def run(self) -> None:
        logger.info(f"--- AVVIO PROCESSO (Dry Run: {self.dry_run}) ---")
        start_t = time.time()
        try:
            self._setup_credentials_and_url()
            lista_ini_pdl = self._initialize_pdl_list()
        except (CriticalConfigError, AutomationException) as e_i:
            logger.critical(f"Errore init: {e_i}", exc_info=True)
            return
        if not lista_ini_pdl:
            logger.info("Nessun PdL da processare. Uscita.")
            return
        idx_proc, res_salvati = self.state_manager.carica_stato()
        res_corr: list[PDLData] = []
        map_ini = {(p.pdl, p.riga_excel_debug): p for p in lista_ini_pdl}
        map_salvati = {(p.pdl, p.riga_excel_debug): p for p in res_salvati}
        for k_pdl, pdl_ini in map_ini.items():
            if k_pdl in map_salvati and map_salvati[k_pdl].stato_script != "Non Processato":
                res_corr.append(map_salvati[k_pdl])
            else:
                res_corr.append(pdl_ini)
        if len(res_corr) != len(lista_ini_pdl):
            logger.warning("Discrepanza stato. Ricreo lista.")
            res_corr = lista_ini_pdl
            idx_proc = -1
        idx_curr_pdl = idx_proc + 1
        setup_ko = 0
        drv = None
        automator = None
        while idx_curr_pdl < len(res_corr):
            pdl_rimanenti = len(res_corr) - idx_curr_pdl
            logger.info(f"Inizio/Ripresa sessione. PdL rimanenti: {pdl_rimanenti} (da idx {idx_curr_pdl}).")
            try:
                if drv is None or automator is None:
                    logger.info("Setup sessione browser...")
                    drv = self.web_driver_manager.get_driver()
                    automator = SafeWorkAutomator(drv, dry_run=self.dry_run)
                    automator.login(self.website_url, self.username, self.password)
                    automator.navigate_to_pdl_booking_page()
                    logger.info("Setup sessione OK.")
                    setup_ko = 0
                for i in range(idx_curr_pdl, len(res_corr)):
                    pdl_c = res_corr[i]
                    if pdl_c.stato_script not in ["Non Processato", "In Elaborazione", None, ""]:
                        logger.info(
                            f"PdL '{pdl_c.pdl}' (Riga {pdl_c.riga_excel_debug}) già processato: '{pdl_c.stato_script}'. Salto."
                        )
                        idx_proc = i
                        idx_curr_pdl = i + 1
                        continue
                    pdl_c.stato_script = "In Elaborazione"
                    self.state_manager.salva_stato(idx_proc, res_corr)
                    try:
                        n_stato = automator.process_single_pdl(pdl_c)
                        pdl_c.stato_script = n_stato
                    except TimeoutAlertDetected as tae:
                        logger.exception(f"Alert timeout PdL '{pdl_c.pdl}': {tae}. Riavvio.")
                        pdl_c.stato_script = "Errore (Timeout Sito)"
                        raise
                    except AutomationException as e_p:
                        logger.error(f"Errore PdL '{pdl_c.pdl}': {e_p}", exc_info=True)
                        pdl_c.stato_script = f"Errore ({type(e_p).__name__})"
                    idx_proc = i
                    self.state_manager.salva_stato(idx_proc, res_corr)
                    idx_curr_pdl = i + 1
                    if i < len(res_corr) - 1:
                        logger.info(f"Pausa {Config.PAUSA_TRA_PDL}s.")
                        time.sleep(Config.PAUSA_TRA_PDL)
                logger.info("Tutti PdL sessione processati.")
                break
            except (KeyboardInterrupt, ProcessInterruptedException) as e_ui:
                logger.info(f"Interrotto: {e_ui}")
                self.state_manager.salva_stato(idx_proc, res_corr)
                raise
            except (TimeoutAlertDetected, AutomationException) as e_sess:
                logger.error(f"Errore sessione: {e_sess}. Riavvio.", exc_info=True)
                setup_ko += 1
                self.state_manager.salva_stato(idx_proc, res_corr)
                if setup_ko >= Config.MAX_SETUP_ATTEMPTS:
                    logger.critical(f"Max ({Config.MAX_SETUP_ATTEMPTS}) tentativi falliti. Interruzione.")
                    break
                logger.info(
                    f"Riavvio sessione tra {Config.PAUSA_TRA_TENTATIVI_SETUP_FALLITI}s (tentativo {setup_ko}/{Config.MAX_SETUP_ATTEMPTS})..."
                )
                self.web_driver_manager.quit_driver()
                drv = None
                automator = None
                time.sleep(Config.PAUSA_TRA_TENTATIVI_SETUP_FALLITI)
            except Exception as e_unexp:
                logger.critical(f"Errore imprevisto: {e_unexp}", exc_info=True)
                setup_ko += 1
                self.state_manager.salva_stato(idx_proc, res_corr)
                if setup_ko >= Config.MAX_SETUP_ATTEMPTS:
                    logger.critical(f"Max ({Config.MAX_SETUP_ATTEMPTS}) tentativi falliti. Interruzione.")
                    break
                self.web_driver_manager.quit_driver()
                drv = None
                automator = None
                time.sleep(Config.PAUSA_TRA_TENTATIVI_SETUP_FALLITI)

        logger.info("Elaborazione PdL terminata.")
        # Logica finale per rimozione file stato
        if idx_curr_pdl >= len(res_corr) and setup_ko < Config.MAX_SETUP_ATTEMPTS:
            logger.info("Processo completato con successo. Rimozione file stato.")
            if not self.dry_run:
                self.state_manager.rimuovi_file_stato()
            else:
                logger.info("[DRY RUN] Rimozione file stato saltata.")
        else:
            logger.warning("Processo interrotto prima della fine. Stato non rimosso.")

        self.web_driver_manager.quit_driver()
        end_t = time.time()
        logger.info(f"--- PROCESSO TERMINATO --- Durata: {end_t - start_t:.2f}s.")


if __name__ == "__main__":
    critical_debug_filepath = os.path.join(ACTUAL_SCRIPT_DIRECTORY, "CRITICAL_EXECUTION_ERROR.txt")
    orchestrator = None
    args = None
    try:
        parser = argparse.ArgumentParser(description="Automatizza prenotazione PdL.")
        parser.add_argument("--dry-run", action="store_true", help="Modalità simulazione.")
        parser.add_argument("--secure-password", action="store_true", help="Richiede PW interattivamente.")
        parser.add_argument("--debug", action="store_true", help="Attiva il logging di livello DEBUG.")
        args = parser.parse_args()
        if args.debug:
            logging.getLogger("PDLAutomator").setLevel(logging.DEBUG)
            for h in logging.getLogger("PDLAutomator").handlers:
                h.setLevel(logging.DEBUG)
            logger.info("Logging DEBUG attivato.")
        if args.secure_password:
            logger.info("Input PW sicuro ATTIVATO.")
        else:
            logger.warning("ATTENZIONE: PW da Excel (NON SICURO). Usa --secure-password.")
        if not WIN32COM_AVAILABLE:
            print(
                "\n"
                + "=" * 80
                + "\n!!! ATTENZIONE: 'pywin32' MANCANTE !!!\n"
                + "Esegui: pip install pywin32\n"
                + "Le funzionalità di macro Excel falliranno.\n"
                + "=" * 80
                + "\n"
            )
        orchestrator = PDLOrchestrator(dry_run=args.dry_run, secure_password_input=args.secure_password)
        orchestrator.run()
    except Exception as e_global:
        with open(critical_debug_filepath, "a", encoding="utf-8") as f_debug:
            f_debug.write(f"--- ERRORE CRITICO: {datetime.now()} ---\n{traceback.format_exc()}\n\n")
        logger.critical(f"ERRORE GLOBALE NON GESTITO: {e_global}", exc_info=True)
    finally:
        if orchestrator and orchestrator.web_driver_manager:
            orchestrator.web_driver_manager.quit_driver()
        logger.info("Script terminato (blocco finally).")
